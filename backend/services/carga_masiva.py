"""
Carga masiva de comprobantes — formato ENERED (compatible con la plantilla oficial de la ATU).

Genera la plantilla Excel ya personalizada con la flota del transportista y la lee de vuelta,
validando cada fila con las reglas del DU 004-2026.

Columnas (mismas que exige la ATU, hoja "Carga"):
  Serie · Numero · Fecha de emision · Fecha inicio periodo · Fecha fin periodo ·
  RUC del grifo · Departamento · Provincia · Distrito · Direccion del grifo ·
  Placa · Categoria · Combustible · Volumen de Galones ·
  Tiene nota de credito · Serie N/C · Numero N/C · Alcance N/C
"""
from __future__ import annotations
import io
import re
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Encabezados exactos que espera la ATU (el orden importa para la compatibilidad).
COLUMNAS = [
    ("serie", "Serie", 10),
    ("numero", "Numero", 14),
    ("fecha_emision", "Fecha de emision", 16),
    ("fecha_inicio", "Fecha inicio periodo", 18),
    ("fecha_fin", "Fecha fin periodo", 18),
    ("ruc_grifo", "RUC del grifo", 15),
    ("departamento", "Departamento", 16),
    ("provincia", "Provincia", 16),
    ("distrito", "Distrito", 16),
    ("direccion_grifo", "Direccion del grifo", 34),
    ("placa", "Placa", 12),
    ("categoria", "Categoria", 11),
    ("combustible", "Combustible", 13),
    ("galones", "Volumen de Galones", 18),
    ("tiene_nc", "Tiene nota de credito", 20),
    ("serie_nc", "Serie N/C", 11),
    ("numero_nc", "Numero N/C", 13),
    ("alcance_nc", "Alcance N/C", 13),
]
GRUPOS = [
    ("Factura Electrónica", 1, 2), ("Fechas", 3, 5), ("Datos del Grifo", 6, 10),
    ("Vehículos Abastecidos", 11, 14), ("Nota de Crédito", 15, 18),
]
CATEGORIAS = ["M2", "M3", "N1", "N2", "N3"]
COMBUSTIBLES = ["B5", "B20"]
_MORADO = "FF7C3AED"
_MORADO_CLARO = "FFF3E8FF"


def _fill(c): return PatternFill("solid", fgColor=c)


def generar_plantilla(*, empresa: str = "", ruc: str = "", vehiculos: list[dict] | None = None) -> bytes:
    """Crea la plantilla ENERED lista para llenar, con la flota del transportista precargada."""
    vehiculos = vehiculos or []
    wb = Workbook()

    # ── Hoja Carga
    ws = wb.active
    ws.title = "Carga"
    ws.freeze_panes = "A4"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS))
    t = ws.cell(1, 1, f"Carga masiva de comprobantes · ENERED{f' — {empresa}' if empresa else ''}")
    t.font = Font(name="Calibri", size=14, bold=True, color="FFFFFFFF")
    t.fill = _fill(_MORADO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for nombre, ini, fin in GRUPOS:  # fila 2: grupos
        ws.merge_cells(start_row=2, start_column=ini, end_row=2, end_column=fin)
        c = ws.cell(2, ini, nombre)
        c.font = Font(size=10, bold=True, color="FF5B21B6")
        c.fill = _fill(_MORADO_CLARO)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    borde = Side(style="thin", color="FFD8B4FE")
    for i, (_, titulo, ancho) in enumerate(COLUMNAS, 1):  # fila 3: encabezados
        c = ws.cell(3, i, titulo)
        c.font = Font(size=10, bold=True, color="FFFFFFFF")
        c.fill = _fill("FFB026FF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=borde)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[3].height = 32
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNAS))}3"

    # Listas desplegables (mismas que la ATU)
    idx = {k: i + 1 for i, (k, _, _) in enumerate(COLUMNAS)}
    FILA_FIN = 2000

    def _dv(formula, col_key, prompt):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                            showDropDown=False, promptTitle="ENERED", prompt=prompt)
        ws.add_data_validation(dv)
        L = get_column_letter(idx[col_key])
        dv.add(f"{L}4:{L}{FILA_FIN}")

    _dv(f'"{",".join(CATEGORIAS)}"', "categoria", "Categoría de la unidad")
    _dv(f'"{",".join(COMBUSTIBLES)}"', "combustible", "B5 o B20")
    _dv('"SI,NO"', "tiene_nc", "¿El comprobante tiene nota de crédito?")
    _dv('"Total,Parcial"', "alcance_nc", "Alcance de la nota de crédito")
    if vehiculos:
        _dv("=Flota!$A$2:$A$" + str(len(vehiculos) + 1), "placa", "Elige una placa de tu flota")

    # ── Hoja Flota (placas del transportista, para el desplegable)
    wf = wb.create_sheet("Flota")
    for j, h in enumerate(["Placa", "Categoria"], 1):
        c = wf.cell(1, j, h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = _fill(_MORADO)
        wf.column_dimensions[get_column_letter(j)].width = 16
    for i, v in enumerate(vehiculos, 2):
        wf.cell(i, 1, (v.get("placa") or "").upper())
        wf.cell(i, 2, (v.get("categoria") or "").upper())

    # ── Hoja Instrucciones
    wi = wb.create_sheet("Instrucciones")
    wi.column_dimensions["A"].width = 30
    wi.column_dimensions["B"].width = 95
    c = wi.cell(1, 1, "Cómo llenar esta plantilla")
    c.font = Font(size=14, bold=True, color="FF5B21B6")
    filas = [
        ("Una fila por placa", "Si un mismo comprobante abastece varias placas, repite la Serie y el Número en cada fila y cambia solo la placa y los galones."),
        ("Serie y Número", "Tal como figuran en la factura (ej. Serie F001, Número 0001234)."),
        ("Fecha de emisión", "Formato dd/mm/aaaa. Debe estar entre 29/05/2026 y 29/07/2026."),
        ("RUC del grifo", "11 dígitos. ENERED verifica solo si está inscrito en OSINERGMIN y completa su ubicación."),
        ("Departamento / Provincia / Distrito / Dirección", "Puedes dejarlos en blanco: ENERED los completa con el RUC del grifo."),
        ("Placa", "Elige una placa de tu flota (lista desplegable). Debe estar registrada en ENERED."),
        ("Categoría", "M2, M3, N1, N2 o N3. Solo estas reciben subsidio."),
        ("Combustible", "B5 o B20."),
        ("Volumen de Galones", "Número con hasta dos decimales (ej. 320.50). No superes el tope de tu categoría."),
        ("Nota de crédito", "Si el comprobante tiene N/C, marca SI y completa serie, número y alcance (Total o Parcial)."),
        ("", ""),
        ("Al subir el archivo", "ENERED valida cada fila (periodo, placa de tu flota, tope de galones, duplicados) y te dice exactamente qué corregir."),
    ]
    for i, (a, b) in enumerate(filas, 3):
        ca = wi.cell(i, 1, a); ca.font = Font(bold=True, size=10)
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        cb = wi.cell(i, 2, b); cb.font = Font(size=10)
        cb.alignment = Alignment(vertical="top", wrap_text=True)

    # ── Pie de marca en Carga
    pie = FILA_FIN + 2
    ws.merge_cells(start_row=pie, start_column=1, end_row=pie, end_column=len(COLUMNAS))
    p = ws.cell(pie, 1, f"ENERED | Subsidio DU 004-2026{f' | RUC {ruc}' if ruc else ''} | Plantilla de carga masiva")
    p.font = Font(size=9, color="FF9CA3AF")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────── Lectura ───────────────────────────────
def _txt(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> Optional[float]:
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


def _fecha(v) -> Optional[str]:
    """Devuelve ISO YYYY-MM-DD desde fecha de Excel o texto dd/mm/aaaa."""
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, (datetime, date)):
        return (v.date() if isinstance(v, datetime) else v).isoformat()
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def leer_plantilla(contenido: bytes) -> list[dict]:
    """Lee la hoja 'Carga' y devuelve una lista de filas normalizadas (con su número de fila)."""
    wb = load_workbook(io.BytesIO(contenido), data_only=True)
    ws = wb["Carga"] if "Carga" in wb.sheetnames else wb.active

    # Localizar la fila de encabezados (por si el archivo viene de la ATU con otro layout).
    fila_hdr = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        vals = [str(c).strip().lower() if c else "" for c in row]
        if "serie" in vals and any("numero" in v for v in vals):
            fila_hdr = i
            break
    if fila_hdr is None:
        raise ValueError("No se encontró la fila de encabezados (Serie, Numero, …).")

    encabezados = [str(c).strip().lower() if c else "" for c in next(
        ws.iter_rows(min_row=fila_hdr, max_row=fila_hdr, values_only=True))]

    def col(*alias) -> Optional[int]:
        for i, h in enumerate(encabezados):
            for a in alias:
                if a in h:
                    return i
        return None

    ix = {
        "serie": col("serie n/c") is not None and col("serie") or col("serie"),
        "numero": col("numero n/c") is not None and col("numero") or col("numero"),
    }
    # Mapeo explícito para no confundir columnas de N/C con las de la factura.
    ix = {
        "serie": next((i for i, h in enumerate(encabezados) if h == "serie"), None),
        "numero": next((i for i, h in enumerate(encabezados) if h == "numero"), None),
        "fecha_emision": col("fecha de emision", "fecha emision"),
        "fecha_inicio": col("fecha inicio"),
        "fecha_fin": col("fecha fin"),
        "ruc_grifo": col("ruc del grifo", "ruc grifo"),
        "departamento": col("departamento"),
        "provincia": col("provincia"),
        "distrito": col("distrito"),
        "direccion_grifo": col("direccion del grifo", "direccion grifo"),
        "placa": col("placa"),
        "categoria": col("categoria"),
        "combustible": col("combustible"),
        "galones": col("volumen de galones", "galones"),
        "tiene_nc": col("tiene nota de credito"),
        "serie_nc": col("serie n/c"),
        "numero_nc": col("numero n/c"),
        "alcance_nc": col("alcance n/c"),
    }

    filas = []
    for n, row in enumerate(ws.iter_rows(min_row=fila_hdr + 1, values_only=True), fila_hdr + 1):
        if not any(c not in (None, "") for c in row):
            continue
        g = lambda k: (row[ix[k]] if ix.get(k) is not None and ix[k] < len(row) else None)
        serie = _txt(g("serie"))
        numero = _txt(g("numero"))
        if not serie and not numero and not _txt(g("placa")):
            continue
        # Ignorar textos que no son datos (p. ej. el pie de marca de la plantilla).
        if serie and not re.fullmatch(r"[A-Za-z]{1,4}\d{0,4}", serie):
            if not _txt(g("placa")) and not numero:
                continue
        if numero:
            numero = re.sub(r"\.0$", "", numero)
        filas.append({
            "fila": n,
            "serie": (serie or "").upper() or None,
            "numero": numero,
            "numero_documento": f"{(serie or '').upper()}-{numero}" if serie and numero else None,
            "fecha": _fecha(g("fecha_emision")),
            "fecha_inicio": _fecha(g("fecha_inicio")),
            "fecha_fin": _fecha(g("fecha_fin")),
            "ruc_emisor": re.sub(r"\D", "", _txt(g("ruc_grifo")) or "") or None,
            "departamento": _txt(g("departamento")),
            "provincia": _txt(g("provincia")),
            "distrito": _txt(g("distrito")),
            "direccion_grifo": _txt(g("direccion_grifo")),
            "placa": (_txt(g("placa")) or "").upper() or None,
            "categoria": (_txt(g("categoria")) or "").upper() or None,
            "producto": (_txt(g("combustible")) or "").upper() or None,
            "galones": _num(g("galones")),
            "tiene_nc": (_txt(g("tiene_nc")) or "").upper() in ("SI", "SÍ", "TRUE", "1"),
            "serie_nc": _txt(g("serie_nc")),
            "numero_nc": _txt(g("numero_nc")),
            "alcance_nc": _txt(g("alcance_nc")),
        })
    return filas
